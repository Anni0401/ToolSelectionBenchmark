#!/usr/bin/env python3
"""Compare WTB in-context and hierarchical runs and write a Markdown report.

The script intentionally uses the generated ``Wild-Tool-Bench_result.jsonl`` files
as the source of truth for correctness, tokens and latency. Selection and tool-call
logs are analysed as supplementary diagnostics.
"""
from __future__ import annotations

import argparse
import json
import math
import statistics
from datetime import datetime, timezone
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable


ROOT = Path(__file__).resolve().parent
DEFAULTS = {
    "data": ROOT / "data/Wild-Tool-Bench.jsonl",
    "in_context_results": ROOT / "result/in_context/langgraph/Wild-Tool-Bench_result.jsonl",
    "hierarchical_results": ROOT / "result/hierarchical/langgraph/Wild-Tool-Bench_result.jsonl",
    "in_context_selection": ROOT / "result/in_context/tool_selection_logs.jsonl",
    "hierarchical_selection": ROOT / "result/hierarchical/tool_selection_logs.jsonl",
    "output": ROOT / "result/in_context_vs_hierarchical_analysis.md",
    "excel": ROOT / "result/wtb_strategy_analysis.xlsx",
}


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as stream:
        for line_no, line in enumerate(stream, 1):
            if not line.strip():
                continue
            try:
                value = json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSON in {path}:{line_no}: {exc}") from exc
            if isinstance(value, dict):
                rows.append(value)
    return rows


def safe_float(value: Any) -> float | None:
    try:
        value = float(value)
        return value if math.isfinite(value) else None
    except (TypeError, ValueError):
        return None


def numeric_total(value: Any, fallback: Any = None) -> float | None:
    """Return a scalar or sum a per-step list stored by WTB."""
    if isinstance(value, list):
        values = [safe_float(item) for item in value]
        values = [item for item in values if item is not None]
        return sum(values) if values else numeric_total(fallback)
    parsed = safe_float(value)
    return parsed if parsed is not None else numeric_total(fallback) if fallback is not None else None


def mean(values: Iterable[float]) -> float | None:
    values = list(values)
    return statistics.mean(values) if values else None


def median(values: Iterable[float]) -> float | None:
    values = list(values)
    return statistics.median(values) if values else None


def pct(numerator: int | float, denominator: int | float) -> str:
    return f"{100 * numerator / denominator:.1f}%" if denominator else "n/a"


def ratio(numerator: int | float, denominator: int | float) -> float | None:
    return numerator / denominator if denominator else None


def fmt(value: Any, digits: int = 2) -> str:
    if value is None:
        return "n/a"
    if isinstance(value, float):
        return f"{value:.{digits}f}"
    return str(value)


def md_escape(value: Any) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ")


def task_index(rows: list[dict[str, Any]]) -> dict[tuple[str, int], dict[str, Any]]:
    output: dict[tuple[str, int], dict[str, Any]] = {}
    for record in rows:
        entry_id = str(record.get("id", ""))
        result = record.get("result", [])
        metrics = record.get("metrics", {}) or {}
        per_task = {int(x["task_idx"]): x for x in metrics.get("per_task", []) if isinstance(x, dict) and "task_idx" in x}
        if not isinstance(result, list):
            result = []
        for fallback_idx, item in enumerate(result):
            if not isinstance(item, dict):
                item = {"action_name_label": "error", "is_optimal": False}
            log = item.get("inference_log", {}) or {}
            idx = log.get("task_idx", fallback_idx)
            try:
                idx = int(idx)
            except (TypeError, ValueError):
                idx = fallback_idx
            metric = per_task.get(idx, {})
            step_count = metric.get("steps")
            if step_count is None:
                step_count = len([k for k in log if str(k).startswith("step_")])
            output[(entry_id, idx)] = {
                "id": entry_id,
                "task_idx": idx,
                "label": item.get("action_name_label", "error"),
                "correct": item.get("action_name_label") == "correct",
                "optimal": bool(item.get("is_optimal", False)),
                "input_tokens": numeric_total(item.get("input_token_count"), metric.get("input_tokens")),
                "output_tokens": numeric_total(item.get("output_token_count"), metric.get("output_tokens")),
                "total_tokens": numeric_total(metric.get("total_tokens")) or (
                    (numeric_total(item.get("input_token_count")) or 0) + (numeric_total(item.get("output_token_count")) or 0)
                ),
                "latency": numeric_total(item.get("latency"), metric.get("llm_latency_s")),
                "wall_time": numeric_total(item.get("task_wall_time_s"), metric.get("task_wall_time_s")),
                "steps": int(step_count or 0),
                "record": item,
            }
    return output


def tool_names_from_task(task: dict[str, Any]) -> list[str]:
    names: list[str] = []
    log = task.get("record", {}).get("inference_log", {}) or {}
    for key in sorted(log, key=lambda x: int(str(x).split("_")[-1]) if str(x).startswith("step_") and str(x).split("_")[-1].isdigit() else -1):
        if not str(key).startswith("step_") or not isinstance(log[key], dict):
            continue
        output = log[key].get("inference_output", {}) or {}
        calls = output.get("tool_calls", []) if isinstance(output, dict) else []
        if not isinstance(calls, list):
            continue
        for call in calls:
            if not isinstance(call, dict):
                continue
            function = call.get("function", {}) or {}
            name = function.get("name") if isinstance(function, dict) else call.get("name")
            if name:
                names.append(str(name))
    return names


def gold_paths(gold: dict[str, Any], key: tuple[str, int]) -> list[list[str]]:
    entry_id, idx = key
    record = gold.get(entry_id, {})
    answers = record.get("english_answer_list", []) or record.get("answer_list", [])
    if idx >= len(answers) or not isinstance(answers[idx], list):
        return []
    # WTB stores one gold action path per task. Alternative paths, where
    # present, are represented by separate answer entries in other datasets;
    # support the current shape without confusing an action with a path.
    task_answer = answers[idx]
    if task_answer and isinstance(task_answer[0], dict) and "action" in task_answer[0]:
        # ``prepare_to_answer`` is a benchmark bookkeeping/finalisation action,
        # not a model tool call, so omit it from tool-call path matching.
        return [[str(x.get("action", {}).get("name")) for x in task_answer
                 if x.get("action", {}).get("name") and x.get("action", {}).get("name") != "prepare_to_answer"]]
    return []


def call_match(predicted: list[str], paths: list[list[str]]) -> tuple[int, bool]:
    if not paths:
        return 0, False
    best = 0
    exact = False
    for gold_path in paths:
        aligned = sum(a == b for a, b in zip(predicted, gold_path))
        best = max(best, aligned)
        exact = exact or predicted == gold_path
    return best, exact


def analyze_selection(rows: list[dict[str, Any]]) -> dict[str, Any]:
    counts = [safe_float(x.get("selected_tools_count")) for x in rows]
    available = [safe_float(x.get("available_tools_count")) for x in rows]
    reductions = [1 - s / a for s, a in zip(counts, available) if s is not None and a]
    return {
        "records": len(rows),
        "selected_mean": mean(x for x in counts if x is not None),
        "selected_median": median(x for x in counts if x is not None),
        "available_mean": mean(x for x in available if x is not None),
        "reduction_mean": mean(reductions),
        "zero_selected": sum(x == 0 for x in counts),
    }


def aggregate(rows: list[dict[str, Any]]) -> dict[str, Any]:
    valid = [x for x in rows.values()]
    return {
        "tasks": len(valid),
        "correct": sum(x["correct"] for x in valid),
        "optimal": sum(x["optimal"] for x in valid),
        "exact_paths": sum(x.get("exact_path", False) for x in valid),
        "predicted_calls": sum(len(x.get("predicted_calls", [])) for x in valid),
        "matched_calls": sum(x.get("matched_calls", 0) for x in valid),
        "gold_calls": sum(len(x.get("gold_paths", [[]])[0]) if x.get("gold_paths") else 0 for x in valid),
        "labels": Counter(x["label"] for x in valid),
        "input_tokens": sum(x["input_tokens"] or 0 for x in valid),
        "output_tokens": sum(x["output_tokens"] or 0 for x in valid),
        "total_tokens": sum(x["total_tokens"] or 0 for x in valid),
        "latency": sum(x["latency"] or 0 for x in valid),
        "wall_time": sum(x["wall_time"] or 0 for x in valid),
        "steps": sum(x["steps"] for x in valid),
        "avg_tokens": mean(x["total_tokens"] for x in valid if x["total_tokens"] is not None),
        "median_tokens": median(x["total_tokens"] for x in valid if x["total_tokens"] is not None),
        "avg_latency": mean(x["latency"] for x in valid if x["latency"] is not None),
        "median_latency": median(x["latency"] for x in valid if x["latency"] is not None),
        "avg_steps": mean(x["steps"] for x in valid),
    }


def discover_result_runs(result_root: Path) -> list[tuple[str, Path, Path]]:
    """Find every result/metrics pair below result_root.

    The folder relative to ``result_root`` is used as the stable run name, so
    rerunning the script updates an existing row instead of duplicating it.
    """
    found = []
    for result_file in sorted(result_root.rglob("Wild-Tool-Bench_result.jsonl")):
        metrics_file = result_file.with_name("Wild-Tool-Bench_metrics_summary.jsonl")
        if metrics_file.exists():
            run_name = result_file.parent.relative_to(result_root).as_posix()
            found.append((run_name, result_file, metrics_file))
    return found


def benchmark_task_index(data_rows: list[dict[str, Any]]) -> dict[tuple[str, int], dict[str, Any]]:
    """Return every task in WTB, including tasks absent from a result file."""
    output: dict[tuple[str, int], dict[str, Any]] = {}
    ordinal = 0
    for record in data_rows:
        entry_id = str(record.get("id", ""))
        tasks = record.get("english_tasks", []) or record.get("tasks", [])
        types = record.get("english_task_types", []) or record.get("task_types", [])
        for task_idx in range(len(tasks)):
            ordinal += 1
            output[(entry_id, task_idx)] = {
                "benchmark_task_number": ordinal,
                "task_type": types[task_idx] if task_idx < len(types) else "unknown",
            }
    return output


def excel_rows(result_file: Path, metrics_file: Path, data_file: Path, run_name: str) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]]]:
    """Create workbook rows using only result and metrics JSONL files."""
    results = task_index(read_jsonl(result_file))
    gold_records = {str(row.get("id")): row for row in read_jsonl(data_file) if row.get("id")}
    benchmark = benchmark_task_index(read_jsonl(data_file))
    task_types: dict[str, list[dict[str, Any]]] = defaultdict(list)
    tasks: list[dict[str, Any]] = []
    for key, benchmark_task in benchmark.items():
        task = results.get(key)
        has_result = task is not None
        if task is None:
            task = {
                "id": key[0], "task_idx": key[1], "label": "no result", "correct": False,
                "optimal": False, "input_tokens": None, "output_tokens": None,
                "total_tokens": None, "latency": None, "wall_time": None, "steps": 0,
                "record": {},
            }
        paths = gold_paths(gold_records, key)
        predicted = tool_names_from_task(task)
        matched, exact = call_match(predicted, paths)
        task.update({"predicted_calls": predicted, "gold_paths": paths, "matched_calls": matched, "exact_path": exact})
        entry_id, task_idx = key
        error_code = has_result and task["label"] == "error"
        correct = has_result and task["correct"]
        # WTB writes a complete result line with ``action_name_label=error``
        # when the model selected the wrong tool. This is therefore both an
        # error-coded result and a not-correct result, not a separate outcome.
        not_correct = has_result and not correct
        task_type = benchmark_task["task_type"]
        row = {
            "Run": run_name, "WTB task number": benchmark_task["benchmark_task_number"],
            "Task ID": entry_id, "Task index": task_idx, "Task reference": f"{entry_id}[{task_idx}]",
            "Task type": task_type, "Result status": "Correct" if correct else "Error code" if error_code else "Not correct" if not_correct else "No result",
            "Returned result": int(has_result), "No result": int(not has_result), "Error code": int(error_code),
            "Not correct": int(not_correct), "Correct": int(correct),
            "Failed task": int(not has_result or error_code),
            "Optimal path": int(task["optimal"]), "Exact tool-call path": int(exact),
            "Correct tool calls": matched, "Total expected tool calls": len(paths[0]) if paths else 0,
            "Predicted tool calls": len(predicted), "Input tokens": task["input_tokens"] or 0,
            "Output tokens": task["output_tokens"] or 0, "Total tokens": task["total_tokens"] or 0,
            "LLM latency (s)": task["latency"] or 0, "Wall time (s)": task["wall_time"] or 0,
            "Steps": task["steps"],
        }
        row["Failed tool calls"] = max(row["Predicted tool calls"] - row["Correct tool calls"], 0)
        row["Failed tool-call rate"] = ratio(row["Failed tool calls"], row["Predicted tool calls"])
        tasks.append(row)
        task_types[task_type].append(row)
    result_metrics = aggregate(results)
    selection_file = result_file.parent.parent / "tool_selection_logs.jsonl"
    selection_metrics = analyze_selection(read_jsonl(selection_file))
    benchmark_tasks = len(tasks)
    returned_tasks = sum(row["Returned result"] for row in tasks)
    no_results = sum(row["No result"] for row in tasks)
    error_codes = sum(row["Error code"] for row in tasks)
    not_correct = sum(row["Not correct"] for row in tasks)
    correct_tasks = sum(row["Correct"] for row in tasks)
    failed_tasks = sum(row["Failed task"] for row in tasks)
    summary = {
        "Run": run_name, "Result file": str(result_file), "Metrics file": str(metrics_file),
        "Tasks": benchmark_tasks, "Evaluated tasks": returned_tasks, "No results": no_results,
        "Error-code results": error_codes, "Not correct results": not_correct,
        "Failed tasks": failed_tasks, "Failure rate": ratio(failed_tasks, benchmark_tasks),
        "No-result rate": ratio(no_results, benchmark_tasks), "Error-code rate": ratio(error_codes, benchmark_tasks),
        "Not-correct rate": ratio(not_correct, benchmark_tasks), "Correct tasks": correct_tasks,
        "Correct task rate": ratio(correct_tasks, benchmark_tasks), "Coverage rate": ratio(returned_tasks, benchmark_tasks),
        "Optimal paths": result_metrics["optimal"], "Optimal path rate": ratio(result_metrics["optimal"], benchmark_tasks),
        "Exact tool-call paths": result_metrics["exact_paths"], "Exact path rate": ratio(result_metrics["exact_paths"], benchmark_tasks),
        "Correct tool calls": result_metrics["matched_calls"], "Correct call rate": ratio(result_metrics["matched_calls"], result_metrics["gold_calls"]),
        "Failed tool calls": max(result_metrics["predicted_calls"] - result_metrics["matched_calls"], 0),
        "Failed tool-call rate": ratio(max(result_metrics["predicted_calls"] - result_metrics["matched_calls"], 0), result_metrics["predicted_calls"]),
        "Total expected tool calls": result_metrics["gold_calls"], "Predicted tool calls": result_metrics["predicted_calls"],
        "Input tokens": result_metrics["input_tokens"], "Output tokens": result_metrics["output_tokens"],
        "Total tokens": result_metrics["total_tokens"], "Avg tokens/task": result_metrics["avg_tokens"],
        "Median tokens/task": result_metrics["median_tokens"], "LLM latency (s)": result_metrics["latency"],
        "Avg latency/task (s)": result_metrics["avg_latency"], "Median latency/task (s)": result_metrics["median_latency"],
        "Wall time (s)": result_metrics["wall_time"], "Tool-call steps": result_metrics["steps"],
        "Avg steps/task": result_metrics["avg_steps"],
        "Selection records": selection_metrics["records"],
        "Avg selected tools/selection": selection_metrics["selected_mean"],
        "Median selected tools/selection": selection_metrics["selected_median"],
        "Avg emitted tool calls/task": ratio(result_metrics["predicted_calls"], result_metrics["tasks"]),
        "Median emitted tool calls/task": median(len(task.get("predicted_calls", [])) for task in results.values()),
    }
    type_rows = []
    for task_type, group in sorted(task_types.items()):
        type_rows.append({
            "Run": run_name, "Task type": task_type, "Tasks": len(group),
            "Evaluated tasks": sum(x["Returned result"] for x in group),
            "No results": sum(x["No result"] for x in group), "Error-code results": sum(x["Error code"] for x in group),
            "Not correct results": sum(x["Not correct"] for x in group),
            "Failed tasks": sum(x["Failed task"] for x in group), "Correct tasks": sum(x["Correct"] for x in group),
            "Optimal paths": sum(x["Optimal path"] for x in group), "Exact tool-call paths": sum(x["Exact tool-call path"] for x in group),
            "Correct tool calls": sum(x["Correct tool calls"] for x in group), "Total expected tool calls": sum(x["Total expected tool calls"] for x in group),
            "Predicted tool calls": sum(x["Predicted tool calls"] for x in group),
            "Input tokens": sum(x["Input tokens"] for x in group), "Output tokens": sum(x["Output tokens"] for x in group),
            "Total tokens": sum(x["Total tokens"] for x in group), "LLM latency (s)": sum(x["LLM latency (s)"] for x in group),
            "Wall time (s)": sum(x["Wall time (s)"] for x in group),
        })
        type_rows[-1].update({
            "Failure rate": ratio(type_rows[-1]["Failed tasks"], type_rows[-1]["Tasks"]),
            "No-result rate": ratio(type_rows[-1]["No results"], type_rows[-1]["Tasks"]),
            "Error-code rate": ratio(type_rows[-1]["Error-code results"], type_rows[-1]["Tasks"]),
            "Not-correct rate": ratio(type_rows[-1]["Not correct results"], type_rows[-1]["Tasks"]),
            "Correct task rate": ratio(type_rows[-1]["Correct tasks"], type_rows[-1]["Tasks"]),
            "Optimal path rate": ratio(type_rows[-1]["Optimal paths"], type_rows[-1]["Tasks"]),
            "Exact path rate": ratio(type_rows[-1]["Exact tool-call paths"], type_rows[-1]["Tasks"]),
            "Correct call rate": ratio(type_rows[-1]["Correct tool calls"], type_rows[-1]["Total expected tool calls"]),
            "Failed tool calls": sum(x["Failed tool calls"] for x in group),
        })
        type_rows[-1]["Failed tool-call rate"] = ratio(type_rows[-1]["Failed tool calls"], type_rows[-1]["Predicted tool calls"])
    return summary, tasks, type_rows


def write_excel(result_root: Path, data_file: Path, output: Path) -> int:
    """Upsert all discovered runs into one Excel workbook."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("Excel-Ausgabe benötigt openpyxl. Installiere es mit: pip install openpyxl") from exc

    sheets = {"Run Summary": [], "Task Details": [], "By Task Type": []}
    headers = {
        "Run Summary": ["Run", "Result file", "Metrics file", "Tasks", "Evaluated tasks", "No results", "Error-code results", "Not correct results", "Failed tasks", "Failure rate", "No-result rate", "Error-code rate", "Not-correct rate", "Correct tasks", "Correct task rate", "Coverage rate", "Optimal paths", "Optimal path rate", "Exact tool-call paths", "Exact path rate", "Correct tool calls", "Correct call rate", "Failed tool calls", "Failed tool-call rate", "Total expected tool calls", "Predicted tool calls", "Input tokens", "Output tokens", "Total tokens", "Avg tokens/task", "Median tokens/task", "LLM latency (s)", "Avg latency/task (s)", "Median latency/task (s)", "Wall time (s)", "Tool-call steps", "Avg steps/task", "Selection records", "Avg selected tools/selection", "Median selected tools/selection", "Avg emitted tool calls/task", "Median emitted tool calls/task"],
        "Task Details": ["Run", "WTB task number", "Task ID", "Task index", "Task reference", "Task type", "Result status", "Returned result", "No result", "Error code", "Not correct", "Correct", "Failed task", "Optimal path", "Exact tool-call path", "Correct tool calls", "Failed tool calls", "Failed tool-call rate", "Total expected tool calls", "Predicted tool calls", "Input tokens", "Output tokens", "Total tokens", "LLM latency (s)", "Wall time (s)", "Steps"],
        "By Task Type": ["Run", "Task type", "Tasks", "Evaluated tasks", "No results", "Error-code results", "Not correct results", "Failed tasks", "Failure rate", "No-result rate", "Error-code rate", "Not-correct rate", "Correct tasks", "Correct task rate", "Optimal paths", "Optimal path rate", "Exact tool-call paths", "Exact path rate", "Correct tool calls", "Correct call rate", "Failed tool calls", "Failed tool-call rate", "Total expected tool calls", "Predicted tool calls", "Input tokens", "Output tokens", "Total tokens", "LLM latency (s)", "Wall time (s)"],
    }
    # Existing rows are retained, but a run with the same folder name is replaced.
    if output.exists():
        workbook = load_workbook(output)
        for sheet_name in headers:
            if sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                existing_headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    existing = dict(zip(existing_headers, row))
                    required_value = {
                        "Run Summary": "Evaluated tasks",
                        "Task Details": "WTB task number",
                        "By Task Type": "Evaluated tasks",
                    }[sheet_name]
                    if existing.get(required_value) is not None:
                        sheets[sheet_name].append(existing)
    else:
        workbook = Workbook()
    discovered = discover_result_runs(result_root)
    new_names = {name for name, _, _ in discovered}
    for sheet_name in sheets:
        key = "Run" if sheet_name == "Run Summary" else "Run"
        sheets[sheet_name] = [row for row in sheets[sheet_name] if row.get(key) not in new_names]
    for run_name, result_file, metrics_file in discovered:
        summary, tasks, type_rows = excel_rows(result_file, metrics_file, data_file, run_name)
        sheets["Run Summary"].append(summary)
        sheets["Task Details"].extend(tasks)
        sheets["By Task Type"].extend(type_rows)
    for sheet_name, columns in headers.items():
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        ws = workbook.create_sheet(sheet_name)
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row in sheets[sheet_name]:
            ws.append([row.get(column) for column in columns])
        for column_index, column in enumerate(columns, 1):
            if "rate" in column.lower():
                for cell in ws.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
                    for value_cell in cell:
                        value_cell.number_format = "0.0%"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        if ws.max_row > 1:
            table = Table(displayName="T" + "".join(c for c in sheet_name if c.isalnum()), ref=ws.dimensions)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)
        for column_cells in ws.columns:
            width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 38)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return len(discovered)


def build_report(args: argparse.Namespace) -> str:
    data_rows = read_jsonl(args.data)
    gold = {str(x.get("id")): x for x in data_rows if x.get("id")}
    strategies = {
        "in_context": task_index(read_jsonl(args.in_context_results)),
        "hierarchical": task_index(read_jsonl(args.hierarchical_results)),
    }
    selections = {
        "in_context": analyze_selection(read_jsonl(args.in_context_selection)),
        "hierarchical": analyze_selection(read_jsonl(args.hierarchical_selection)),
    }
    common = sorted(set(strategies["in_context"]) & set(strategies["hierarchical"]))
    for strategy in strategies:
        for key, task in strategies[strategy].items():
            paths = gold_paths(gold, key)
            predicted = tool_names_from_task(task)
            task["predicted_calls"] = predicted
            task["gold_paths"] = paths
            task["matched_calls"], task["exact_path"] = call_match(predicted, paths)
            task["gold_path_count"] = len(paths)
    aggs = {name: aggregate(tasks) for name, tasks in strategies.items()}

    lines = [
        "# WTB: In-context vs. hierarchical tool selection",
        "",
        "> Automatisch erzeugter Vergleich der vorhandenen WTB-Läufe. Korrektheit und Token-/Latenzwerte stammen aus den `Wild-Tool-Bench_result.jsonl`-Dateien; Selection-Logs werden separat ausgewertet.",
        "",
        "## Executive Summary",
        "",
        f"- Vergleichbare Tasks: **{len(common)}** (In-context insgesamt: {aggs['in_context']['tasks']}, hierarchical insgesamt: {aggs['hierarchical']['tasks']}).",
        f"- Korrekte Tasks: **{aggs['in_context']['correct']}/{aggs['in_context']['tasks']} ({pct(aggs['in_context']['correct'], aggs['in_context']['tasks'])})** vs. **{aggs['hierarchical']['correct']}/{aggs['hierarchical']['tasks']} ({pct(aggs['hierarchical']['correct'], aggs['hierarchical']['tasks'])})**.",
        f"- Optimale Tasks: **{aggs['in_context']['optimal']}/{aggs['in_context']['tasks']} ({pct(aggs['in_context']['optimal'], aggs['in_context']['tasks'])})** vs. **{aggs['hierarchical']['optimal']}/{aggs['hierarchical']['tasks']} ({pct(aggs['hierarchical']['optimal'], aggs['hierarchical']['tasks'])})**.",
        "- `optimal` ist die vom WTB-Evaluator markierte optimale Aktionsauswahl; es ist nicht dasselbe wie eine bloß syntaktisch gültige Tool-Call-Antwort.",
        "",
        "## Gesamtmetriken",
        "",
        "| Metrik | In-context | Hierarchical | Differenz (hier. − in-context) |",
        "|---|---:|---:|---:|",
    ]
    metric_rows = [
        ("Tasks", "tasks", 0), ("Correct tasks", "correct", 0), ("Optimal paths/tasks", "optimal", 0), ("Exact tool-call paths", "exact_paths", 0),
        ("Predicted tool calls", "predicted_calls", 0), ("Gold tool calls", "gold_calls", 0), ("Positionally matched calls", "matched_calls", 0),
        ("Input tokens", "input_tokens", 0), ("Output tokens", "output_tokens", 0), ("Total tokens", "total_tokens", 0),
        ("Ø tokens / task", "avg_tokens", 2), ("Median tokens / task", "median_tokens", 2),
        ("LLM latency (s)", "latency", 2), ("Ø latency / task (s)", "avg_latency", 2), ("Median latency / task (s)", "median_latency", 2),
        ("Wall time (s)", "wall_time", 2), ("Tool-call steps", "steps", 0), ("Ø steps / task", "avg_steps", 2),
    ]
    for label, key, digits in metric_rows:
        left, right = aggs["in_context"][key], aggs["hierarchical"][key]
        diff = right - left if left is not None and right is not None else None
        lines.append(f"| {label} | {fmt(left, digits)} | {fmt(right, digits)} | {fmt(diff, digits)} |")
    lines += [
        "",
        "### Selection-Overhead",
        "",
        "| Metrik | In-context | Hierarchical |",
        "|---|---:|---:|",
    ]
    for label, key, digits in [("Selection log records", "records", 0), ("Ø available tools", "available_mean", 1), ("Ø selected tools", "selected_mean", 1), ("Median selected tools", "selected_median", 1), ("Ø reduction", "reduction_mean", 3), ("Empty selections", "zero_selected", 0)]:
        suffix = "" if key != "reduction_mean" else " (als Anteil)"
        lines.append(f"| {label}{suffix} | {fmt(selections['in_context'][key], digits)} | {fmt(selections['hierarchical'][key], digits)} |" if key != "reduction_mean" else f"| {label} | {pct(selections['in_context'][key], 1)} | {pct(selections['hierarchical'][key], 1)} |")
    lines += [
        "",
        "## Paarweiser Task-Vergleich",
        "",
        "| Ergebnis | Anzahl | Anteil der gemeinsamen Tasks |",
        "|---|---:|---:|",
    ]
    outcomes = Counter()
    for key in common:
        a, h = strategies["in_context"][key], strategies["hierarchical"][key]
        outcomes["Beide korrekt" if a["correct"] and h["correct"] else "Nur in-context korrekt" if a["correct"] else "Nur hierarchical korrekt" if h["correct"] else "Beide inkorrekt"] += 1
    for label in ["Beide korrekt", "Nur in-context korrekt", "Nur hierarchical korrekt", "Beide inkorrekt"]:
        lines.append(f"| {label} | {outcomes[label]} | {pct(outcomes[label], len(common))} |")
    lines += ["", "### Tasks mit deutlichem Unterschied", "", "Die folgenden Tabellen listen die paarweisen Fälle, in denen genau eine Strategie korrekt bzw. optimal war.", "", "| Task | In-context | Hierarchical | Text (gekürzt) |", "|---|---|---|---|"]
    task_text = {str(x.get("id")): x for x in data_rows if x.get("id")}
    diffs = []
    for key in common:
        a, h = strategies["in_context"][key], strategies["hierarchical"][key]
        if a["correct"] != h["correct"] or a["optimal"] != h["optimal"]:
            entry, idx = key
            source = task_text.get(entry, {})
            texts = source.get("english_tasks", [])
            text = texts[idx] if idx < len(texts) else ""
            diffs.append((key, a, h, text))
            lines.append(f"| {entry} / {idx} | {a['label']} / optimal={a['optimal']} | {h['label']} / optimal={h['optimal']} | {md_escape(text[:180])} |")
    if not diffs:
        lines.append("| — | Keine abweichenden Tasks im gemeinsamen Datensatz | — | — |")
    lines += ["", "## Gemeinsame Schwächen", "", "| Schwäche | In-context | Hierarchical |", "|---|---:|---:|"]
    both_errors = [key for key in common if not strategies["in_context"][key]["correct"] and not strategies["hierarchical"][key]["correct"]]
    for label, value in [("Beide inkorrekt", len(both_errors)), ("Beide nicht optimal", sum(not strategies['in_context'][k]['optimal'] and not strategies['hierarchical'][k]['optimal'] for k in common)), ("Beide mit mindestens einem Tool-Call", sum(bool(strategies['in_context'][k]['predicted_calls']) and bool(strategies['hierarchical'][k]['predicted_calls']) for k in common))]:
        lines.append(f"| {label} | {value if label != 'Beide mit mindestens einem Tool-Call' else '—'} | {pct(value, len(common))} |")
    lines += ["", "### Fehler nach Task-Typ", "", "| Task-Typ | Gemeinsame Tasks | Beide inkorrekt | In-context Fehler | Hierarchical Fehler |", "|---|---:|---:|---:|---:|"]
    by_type: dict[str, list[tuple[str, int]]] = defaultdict(list)
    for key in common:
        entry, idx = key
        record = task_text.get(entry, {})
        types = record.get("english_task_types", [])
        by_type[types[idx] if idx < len(types) else "unknown"].append(key)
    for typ, keys in sorted(by_type.items()):
        both = sum(not strategies['in_context'][k]['correct'] and not strategies['hierarchical'][k]['correct'] for k in keys)
        ia = sum(not strategies['in_context'][k]['correct'] for k in keys)
        hh = sum(not strategies['hierarchical'][k]['correct'] for k in keys)
        lines.append(f"| {md_escape(typ)} | {len(keys)} | {both} ({pct(both,len(keys))}) | {ia} ({pct(ia,len(keys))}) | {hh} ({pct(hh,len(keys))}) |")
    lines += ["", "## Methodik und Einschränkungen", "", "- Die Runs sind nicht vollständig gleich groß: Es werden nur Tasks verglichen, die in beiden Result-Dateien vorhanden sind.", "- Die in-context-Tool-Call-Logs enthalten keine `test_entry_id`/`task_idx`; deshalb werden Call-Details primär aus den Result-Inference-Logs gelesen und Selection-Logs nur aggregiert.", "- `matched tool calls` werden gegen die WTB-Gold-Actionpfade positionsweise gematcht; bei mehreren Goldpfaden wird das beste Matching verwendet.", "- Ein hoher Tokenwert im In-context-Lauf ist erwartbar, weil dort sehr viele Tools an das ausführende LLM übergeben werden. Das ist Overhead, aber nicht automatisch ein Qualitätsfehler.", ""]
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    for name, default in DEFAULTS.items():
        flag = "--" + name.replace("_", "-")
        parser.add_argument(flag, type=Path, default=default)
    args = parser.parse_args()
    report = build_report(args)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(report, encoding="utf-8")
    print(f"Wrote {args.output}")
    run_count = write_excel(args.output.parent, args.data, args.excel)
    print(f"Wrote {args.excel} ({run_count} result folders; existing matching folders were updated)")


if __name__ == "__main__":
    main()
